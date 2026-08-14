"""One-time import of historical data from the old '2026-27 Budgeting.xlsx' workbook.

Usage:
    python import_xlsx.py "C:\\path\\to\\2026-27 Budgeting.xlsx"

Imports:
  - Transactions sheet -> expense transactions
  - Income sheet       -> income transactions
  - Savings sheet      -> savings goals (using the sheet's "Current Amount" as each
                           goal's starting balance, dated today, so progress bars are
                           accurate immediately without double-counting past contributions)
  - Budget Tracking     -> per-category budgets (skips the aggregate "Bills" row, which
    sheet                 bundled 5 categories together in the old sheet; set those
                           individually on the Budget page if you want them)
"""
from __future__ import annotations

import sys
import datetime as dt

import openpyxl

import db


def import_workbook(path: str) -> None:
    db.init_db()
    wb = openpyxl.load_workbook(path, data_only=True)

    # ---- Transactions (expenses) ----
    added_expenses = 0
    ws = wb["Transactions"]
    for row in ws.iter_rows(min_row=2, values_only=False):
        date_cell, amount_cell, category_cell, desc_cell = row[0], row[1], row[2], row[3]
        if date_cell.value is None or amount_cell.value is None or category_cell.value is None:
            continue
        date_val = date_cell.value
        date_str = date_val.date().isoformat() if isinstance(date_val, dt.datetime) else str(date_val)
        db.add_transaction(date_str, "expense", str(category_cell.value), desc_cell.value or "", float(amount_cell.value))
        added_expenses += 1

    # ---- Income ----
    added_income = 0
    ws = wb["Income"]
    for row in ws.iter_rows(min_row=2, values_only=False):
        date_cell, amount_cell, category_cell, desc_cell = row[0], row[1], row[2], row[3]
        if date_cell.value is None or amount_cell.value is None or category_cell.value is None:
            continue
        date_val = date_cell.value
        date_str = date_val.date().isoformat() if isinstance(date_val, dt.datetime) else str(date_val)
        db.add_transaction(date_str, "income", str(category_cell.value), desc_cell.value or "", float(amount_cell.value))
        added_income += 1

    # ---- Savings goals ----
    added_goals = 0
    ws = wb["Savings"]
    today = dt.date.today().isoformat()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        name, monthly_plan, _invested_since, _starting_month_amount, goal_amount, current_amount = row[:6]
        if not name:
            continue
        db.add_savings_goal(
            str(name), float(goal_amount or 0), float(monthly_plan or 0), float(current_amount or 0), today
        )
        added_goals += 1

    # ---- Budgets (skip aggregate rows like "Bills" that bundle several categories) ----
    added_budgets = 0
    skipped = []
    ws = wb["Budget Tracking"]
    known_categories = set(db.EXPENSE_CATEGORIES)
    for row in ws.iter_rows(min_row=2, values_only=True):
        category, budget = row[0], row[1]
        if not category or budget is None:
            continue
        if category not in known_categories:
            skipped.append(str(category))
            continue
        db.set_budget(str(category), float(budget))
        added_budgets += 1

    print(f"Imported {added_expenses} expense transactions")
    print(f"Imported {added_income} income transactions")
    print(f"Imported {added_goals} savings goals")
    print(f"Imported {added_budgets} budgets")
    if skipped:
        print(f"Skipped budget row(s) that don't match a known category (set these manually): {skipped}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python import_xlsx.py <path to xlsx>")
        sys.exit(1)
    import_workbook(sys.argv[1])
